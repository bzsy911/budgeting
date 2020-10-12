import os
import numpy as np
import pandas as pd
import openpyxl
from source.base import CitiBudgeting


class Validator(CitiBudgeting):

    def __init__(self, excel, raw_tab='TimeAndExpenseDetails', staffing_tab='Tech', bill_tab='Bill'):
        super().__init__(excel)
        self.name = 'report'
        self.names = self._process_names(bill_tab)
        # self.refresh_log = self._refresh_raw(raw_tab)
        self.ts_df, self.ts_summary = self._process_raw(raw_tab)
        self.staffing_df, self.staffing_summary = self._process_staffing(staffing_tab)
        self.total = None
        self.functions = [self._get_summary, self._get_charging_diffs]

    def run_report(self, save=True):
        super().run_report_txt(self.functions, save)

    def run_comparison(self, old_file, save=True):
        if self.total is not None:
            self._get_summary()
        old = Validator(old_file)
        old._get_summary()
        diff = (self.total.subtract(old.total, fill_value=0)).applymap(lambda x: np.nan if x == 0 else x).dropna(
            how='all').dropna(how='all', axis=1).fillna('')
        msg = f"""
#### Comparison with Old Spreedsheet #### 
{self.excel}
{self.total}

{old.excel}
{old.total}

Delta
{diff}
"""
        print(msg)
        if save:
            file_name = f'{self.output}Comparison {self.excel.split()[-1][:4]} and {old.excel.split()[-1][:4]}.txt'
            with open(file_name, 'w+') as f:
                f.write(msg)
            print(f'Result has been saved to {file_name}')
        return msg

    # ---------- App Functions ----------
    def _refresh_raw(self, raw_tab):
        # Currently not in use because of bug.
        # Bug: after rewriting, other tabs are affected. pd.read_excel header has issues. Can't call by value.
        time_sheets = os.listdir(self.input + self.time_exp)
        if time_sheets:
            file_name = sorted(time_sheets)[-1]

            ts_wb = openpyxl.load_workbook(filename=self.input+self.time_exp+file_name)
            ws = ts_wb.worksheets[0]

            tracker_wb = openpyxl.load_workbook(filename=self.excel)
            tracker_wb.remove(tracker_wb.get_sheet_by_name(raw_tab))
            new_tab = tracker_wb.create_sheet(raw_tab)

            for row in ws:
                for cell in row:
                    if cell.row > 4:
                        new_tab[f'{cell.column_letter}{cell.row-4}'].value = cell.value
            tracker_wb.save(self.excel)

            res = f"Refreshed the {raw_tab} tab by {file_name}.\n\n"
        else:
            res = f"No files found in {self.time_exp} folder. Proceed without refresh.\n\n"
        return res

    def _get_summary(self):
        min_week = str(min(self.ts_summary.keys()))[:10]
        max_week = str(max(self.ts_summary.keys()))[:10]

        total = pd.pivot_table(self.ts_df, values='Charged Hours',
                               index=['Activity Code Description'],
                               columns=['Week Ending Date'],
                               aggfunc=np.sum).fillna(0)
        total.columns = [str(x)[:10] for x in total.columns]

        budget = self.staffing_df.iloc[:-1, :len(total.columns) + 2].groupby('Activity Code').sum()
        diff = (budget - total).applymap(lambda x: min(x, 0)).applymap(lambda x: np.nan if x == 0 else x). \
            dropna(how='all').dropna(how='all', axis=1).fillna('')
        if len(diff) > 0:
            msg = f'Attention: Found unbudgeted hours charged:\n{diff}'
        else:
            msg = f'Great! All charged hours are under budget!'

        total.loc['Total'] = total.sum()
        total['Total'] = total.sum(axis=1)
        budget.loc['Total'] = budget.sum()
        budget['Total'] = budget.sum(axis=1)

        pd.set_option('display.max_columns', None, 'display.expand_frame_repr', False)
        res = f"""
#### 1. SUMMARY ####
Currently {len(self.ts_summary)} weeks of timesheets have been recorded (from {min_week} to {max_week}). 

Total budget hours by Activity Codes by week:
{budget}

Total charged hours by Activity Codes by week:
{total}

{msg}
"""
        self.total = total
        return res

    def _get_charging_diffs(self):
        charged = self.ts_summary
        budget = self.staffing_summary
        over_charge_list = []
        under_charge_list = []
        wrong_code_list = []
        for date in charged:
            over_charge = []
            under_charge = []
            wrong_code = []
            for name in set(charged[date].keys()).union(set(budget[date].keys())):
                if name in budget[date] and name in charged[date]:
                    diff = {}
                    for ws in set(charged[date][name].keys()).union(set(budget[date][name].keys())):
                        diff[ws] = charged[date][name].get(ws, 0) - budget[date][name].get(ws, 0)
                    over_charge.extend(
                        [(name, ws, budget[date][name].get(ws, 0), charged[date][name][ws]) for ws, hrs in
                         diff.items() if hrs > 0])
                    under_charge.extend(
                        [(name, ws, budget[date][name][ws], charged[date][name].get(ws, 0)) for ws, hrs in
                         diff.items() if hrs < 0])
                elif name in budget[date]:
                    under_charge.extend(
                        [(name, ws, budget[date][name][ws], 0) for ws, hrs in budget[date][name].items()])
                else:
                    over_charge.extend(
                        [(name, ws, 0, charged[date][name][ws]) for ws, hrs in charged[date][name].items()])

                if name in [r[0] for r in over_charge] and name in [r[0] for r in under_charge]:
                    over = [r for r in over_charge if r[0] == name]
                    under = [r for r in under_charge if r[0] == name]
                    wrong_code.extend(
                        [f'    {name} charged {r[3]} hours on {r[1]}, with {r[2]} hours on budget.' for r in over])
                    wrong_code.extend(
                        [f'    {name} charged {r[3]} hours on {r[1]}, with {r[2]} hours on budget.' for r in under])

            if over_charge:
                over_charge_list.append(f'\nWeek {date}')
                over_charge_list.extend(
                    [f'    {r[0]} charged {r[3]} hours on {r[1]}, with {r[2]} hours on budget.' for r in
                     over_charge])
            if under_charge:
                under_charge_list.append(f'\nWeek {date}')
                under_charge_list.extend(
                    [f'    {r[0]} charged {r[3]} hours on {r[1]}, with {r[2]} hours on budget.' for r in
                     under_charge])
            if wrong_code:
                wrong_code_list.append(f'\nWeek {date}')
                wrong_code_list.extend(wrong_code)

        if over_charge_list:
            msg = '\n'.join(over_charge_list)
            over_charge_msg = f'Attention: Found over-charging hours! Please consider correct them:\n{msg}'
        else:
            over_charge_msg = f'Great! No over-charging error is found!'
        if under_charge_list:
            msg = '\n'.join(under_charge_list)
            under_charge_msg = f'There are some under-charging hours you may want to notice.\n{msg}'
        else:
            under_charge_msg = f'Great! No under-charging error is found!'
        if wrong_code_list:
            msg = '\n'.join(wrong_code_list)
            wrong_code_msg = f'The following discrepancies may caused by miss use of Activity Code:\n{msg}'
        else:
            wrong_code_msg = f'Great! No activity code is wrongly charged!'
        res = f"""
#### 2. OVER-CHARGING ####
{over_charge_msg}

#### 3. UNDER-CHARGING ####
{under_charge_msg}

#### 4. POSSIBLY WRONG CODE ####
{wrong_code_msg}
"""
        return res

    # ---------- Helper Functions ----------
    def _process_raw(self, raw_tab):
        df = pd.read_excel(self.excel, sheet_name=raw_tab)
        df = df.groupby(['Week Ending Date', 'Employee Name', 'Activity Code Description']).sum().reset_index()
        df = df[df['Charged Hours'] != 0].iloc[:, [0, 1, 2, 6]]
        df = df[df['Week Ending Date'] >= '2020-03-06']
        df['Name'] = df.apply(lambda x: self._get_name(x['Employee Name']), axis=1)

        summary = {}
        for date in df['Week Ending Date'].unique():
            summary[str(date)[:10]] = {}
            for r in df[df['Week Ending Date'] == date].to_dict('records'):
                if r['Name'] in summary[str(date)[:10]]:
                    summary[str(date)[:10]][r['Name']][r['Activity Code Description']] = r['Charged Hours']
                else:
                    summary[str(date)[:10]][r['Name']] = {r['Activity Code Description']: r['Charged Hours']}
        return df, summary

    def _process_names(self, bill_tab):
        df = pd.read_excel(self.excel, sheet_name=bill_tab, header=3)[['Name', 'Legal Name']].drop_duplicates().dropna(how='all')
        return df

    def _get_name(self, legal_name):
        df = self.names[self.names['Legal Name'] == legal_name]['Name'].reset_index()
        if df.shape[0] == 0:
            print(f'WARNING: {legal_name} doesn\'t have a preferred name in record. Processed as Legal Name instead.')
            return legal_name
        return df.at[0, 'Name']

    def _process_staffing(self, staffing_tab):
        df_l = pd.read_excel(self.excel, sheet_name=staffing_tab, header=2)[['Activity Code', 'Name']]\
            .dropna(how='all').iloc[:-1, :].fillna('No Name')
        df_r = pd.read_excel(self.excel, sheet_name=staffing_tab, header=1).iloc[1:df_l.shape[0]+1, 18:]\
            .reset_index(drop=True).fillna(0)
        df_r.columns = [str(x)[:10] for x in df_r.columns]
        df = pd.concat([df_l, df_r], axis=1)
        summary = df_r.to_dict()
        keys = {i: [d['Name'], d['Activity Code']] for i, d in enumerate(df_l.to_dict('records'))}
        for date in summary:
            temp = {}
            for i in summary[date]:
                if summary[date][i] == 0:
                    continue
                if keys[i][0] in temp:
                    temp[keys[i][0]][keys[i][1]] = summary[date][i]
                else:
                    temp[keys[i][0]] = {keys[i][1]: summary[date][i]}
            summary[date] = temp
        return df, summary


if __name__ == "__main__":
    pass
